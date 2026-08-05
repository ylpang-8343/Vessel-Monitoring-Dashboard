import io
from datetime import datetime, timezone

from openpyxl import load_workbook

from app.models import EventType, StatusEvent, Vessel
from app.services.report_service import build_excel_report, build_pdf_report, build_report_summary


def _vessel_with_event(db_session, imo, event_type, destination="Pasir Gudang"):
    vessel = Vessel(name=f"MV {imo}", imo_number=imo, destination_port=destination)
    db_session.add(vessel)
    db_session.commit()
    db_session.refresh(vessel)

    db_session.add(
        StatusEvent(
            vessel_id=vessel.id,
            event_type=event_type,
            current_location=destination or "South China Sea",
            last_event_text=f"Test event {imo}",
            source_name="Mock Tracking Feed",
            occurred_at=datetime.now(timezone.utc),
        )
    )
    db_session.commit()
    return vessel


def test_build_report_summary_categorizes_by_latest_status(db_session):
    _vessel_with_event(db_session, "1111111", EventType.SAILING, destination=None)
    _vessel_with_event(db_session, "2222222", EventType.ETA_DESTINATION)
    _vessel_with_event(db_session, "3333333", EventType.ARRIVED_DESTINATION)
    _vessel_with_event(db_session, "4444444", EventType.AT_PORT)

    summary = build_report_summary(db_session)

    # "active" includes every non-archived vessel regardless of status.
    assert {v.imo_number for v in summary.active} == {"1111111", "2222222", "3333333", "4444444"}
    assert {v.imo_number for v in summary.eta_to_destination} == {"2222222"}
    assert {v.imo_number for v in summary.arrived_at_destination} == {"3333333"}


def test_build_report_summary_excludes_archived_vessels(db_session):
    vessel = _vessel_with_event(db_session, "5555555", EventType.ARRIVED_DESTINATION)
    vessel.archived_at = datetime.now(timezone.utc)
    db_session.commit()

    summary = build_report_summary(db_session)
    assert summary.active == []
    assert summary.arrived_at_destination == []


def test_build_excel_report_has_one_sheet_per_category_with_correct_rows(db_session):
    _vessel_with_event(db_session, "2222222", EventType.ETA_DESTINATION)
    summary = build_report_summary(db_session)

    content = build_excel_report(summary)
    assert content[:2] == b"PK"  # .xlsx files are zip archives

    wb = load_workbook(io.BytesIO(content))
    assert wb.sheetnames == ["Active Vessels", "ETA to Destination", "Arrived at Destination"]

    eta_sheet = wb["ETA to Destination"]
    header = [cell.value for cell in eta_sheet[1]]
    assert header == ["Vessel Name", "IMO Number", "Current Location", "Last Event", "Destination", "Source"]
    data_row = [cell.value for cell in eta_sheet[2]]
    assert data_row[1] == "2222222"

    # Nothing arrived, so that sheet should have only the header row.
    assert wb["Arrived at Destination"].max_row == 1


def test_build_pdf_report_produces_a_valid_pdf(db_session):
    _vessel_with_event(db_session, "3333333", EventType.ARRIVED_DESTINATION)
    summary = build_report_summary(db_session)

    content = build_pdf_report(summary)
    assert content.startswith(b"%PDF")
    assert len(content) > 100


def test_build_pdf_report_handles_empty_categories(db_session):
    # No vessels at all - every section falls into the "No vessels in this category" branch;
    # this should still produce a well-formed (non-empty, valid-header) PDF, not crash.
    summary = build_report_summary(db_session)
    content = build_pdf_report(summary)
    assert content.startswith(b"%PDF")
