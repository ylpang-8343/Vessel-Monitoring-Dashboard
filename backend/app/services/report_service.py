"""Reports (Section 9 Phase 4 / Section 7's "generates daily reports covering active vessels,
ETA to destination, delayed vessels, and arrived vessels, exportable to Excel or PDF").

"Delayed vessels" - the proposal's fourth category - is intentionally not included. Section 3.10
already explains why the dashboard itself never shows a "Delayed" status: AIS-style position
reports don't reliably indicate a delay without a planned ETA to compare against, which this app
doesn't capture anywhere. Guessing at one for a report would produce numbers that look precise
but aren't backed by real data - the same trade-off Section 3.10 already made for the live
dashboard. Real delay detection is explicitly Phase 6 (AI Delay Detection), once there's an
actual predicted-ETA source to compare against.
"""

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from sqlalchemy.orm import Session

from app.models import EventType, Vessel
from app.schemas import ReportSummaryOut, VesselOut
from app.services.presentation import to_vessel_out

# Column order shared by both the Excel and PDF renderers, and matching the dashboard table's
# own column order (Section 3.4) so the report reads consistently with the live view.
COLUMNS = ["Vessel Name", "IMO Number", "Current Location", "Last Event", "Destination", "Source"]


def _row(v: VesselOut) -> list[str]:
    return [
        v.name,
        v.imo_number,
        v.current_location or "—",
        v.last_event_text or "Awaiting first tracking update",
        v.destination_port or "—",
        v.source_name or "—",
    ]


def build_report_summary(db: Session) -> ReportSummaryOut:
    """Gather the three vessel categories, all scoped to *active* (non-archived) vessels - an
    archived vessel is, by definition, no longer part of any of these live operational views."""
    active_vessels = db.query(Vessel).filter(Vessel.archived_at.is_(None)).order_by(Vessel.name).all()
    active = [to_vessel_out(v) for v in active_vessels]
    eta_to_destination = [v for v in active if v.last_event_type == EventType.ETA_DESTINATION]
    arrived_at_destination = [v for v in active if v.last_event_type == EventType.ARRIVED_DESTINATION]

    return ReportSummaryOut(
        active=active,
        eta_to_destination=eta_to_destination,
        arrived_at_destination=arrived_at_destination,
        generated_at=datetime.now(timezone.utc),
    )


def build_excel_report(summary: ReportSummaryOut) -> bytes:
    """Render the summary as an .xlsx workbook, one sheet per category (Section 9 Phase 4's
    "exportable to Excel" requirement)."""
    wb = Workbook()
    # openpyxl always creates one default sheet - reuse it for the first section instead of
    # leaving a stray blank "Sheet" behind.
    sections = [
        ("Active Vessels", summary.active),
        ("ETA to Destination", summary.eta_to_destination),
        ("Arrived at Destination", summary.arrived_at_destination),
    ]
    for i, (title, vessels) in enumerate(sections):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = title[:31]  # Excel sheet names are capped at 31 characters
        ws.append(COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for v in vessels:
            ws.append(_row(v))
        for col in ws.columns:
            width = max(len(str(c.value)) for c in col if c.value is not None) if col else 10
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_pdf_report(summary: ReportSummaryOut) -> bytes:
    """Render the summary as a PDF (Section 9 Phase 4's "exportable to ... PDF" requirement),
    one table per category on a landscape letter page (wide enough for six columns of vessel
    data without truncation)."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Vessel Monitoring Dashboard — Report", styles["Title"]),
        Paragraph(
            f"Generated {summary.generated_at.strftime('%d %b %Y, %H:%M UTC')}",
            styles["Normal"],
        ),
        Spacer(1, 16),
    ]

    sections = [
        ("Active Vessels", summary.active),
        ("ETA to Destination", summary.eta_to_destination),
        ("Arrived at Destination", summary.arrived_at_destination),
    ]
    for title, vessels in sections:
        elements.append(Paragraph(f"{title} ({len(vessels)})", styles["Heading2"]))
        if vessels:
            data = [COLUMNS] + [_row(v) for v in vessels]
            table = Table(data, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0b3d5c")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f4f5")]),
                    ]
                )
            )
            elements.append(table)
        else:
            elements.append(Paragraph("No vessels in this category.", styles["Normal"]))
        elements.append(Spacer(1, 20))

    doc.build(elements)
    return buffer.getvalue()
